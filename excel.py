#-*-coding:utf-8-*-
# Time:2019/1/5 17:46

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.chart import LineChart, Series, Reference

import os

usFreqSendpower = [31]
data = ['pass', 'fail']
data1 = [32]
data2 = [['11223344', '2F2F2F3D3D3D'], ['11223348', '2F2F2F3D3D3D'],
         ['1122334c', '2F2F2F3D3D3D']]


def str_to_hex(s):
    j = hex(int(str(s), 16) + 4).replace('0x', '').strip()
    print(j)
    return j


def readExcel_GetWriteData(excelName):
    # 获取文件路径
    excelPath = os.path.join(os.getcwd())
    #print(excelPath)
    ExcelFullName = os.path.join(excelPath, excelName)
    print(ExcelFullName)

    wb = load_workbook(ExcelFullName)
    #wb = load_workbook(filename=ExcelFullName)

    # 获取当前活跃的worksheet,默认就是第一个worksheet
    #ws = wb.active
    # 当然也可以使用下面的方法
    # 获取所有表格(worksheet)的名字
    sheets = wb.sheetnames
    print(sheets)
    # # 第一个表格的名称
    sheet_first = sheets[0]
    # # 获取特定的worksheet
    #
    ws = wb[sheet_first]
    #print("***")
    #print(sheet_first)
    #print(ws.title)
    #print("^^^")
    # 获取表格所有行和列，两者都是可迭代的
    rows = ws.rows
    #print(rows)

    columns = ws.columns
    # 迭代所有的行
    i = 0
    data = []
    for row in rows:
        line = [col.value for col in row]
        #print(line)
        if (i > 1) and (line[1] != None) and (line[2] != None):
            filterItem = []
            filterItem.append(str(line[1]))
            filterItem.append(str(line[2]))
            data.append(filterItem)
        i += 1
    print(data)
    if data == []:
        print('文件为空，请导出模板文件，并设置要写入的寄存器地址和值，再次导入')
        return -1
    else:
        return data
    # 通过坐标读取值
    #print(ws['A1'].value)  # A表示列,1表示行
    #print(ws.cell(row=1, column=1).value)


def writeAddrValueToExcel(excelName, sheet, valueArray):
    # 获取文件路径
    excelPath = os.path.join(os.getcwd())
    #print(excelPath)
    ExcelFullName = os.path.join(excelPath, excelName)
    print(ExcelFullName)
    try:
        wb = load_workbook(ExcelFullName)
        ws = wb[sheet]
        rows = ws.rows
        #print(rows)
        columns = ws.columns
        # 迭代所有的行
        i = 1
        j = 0
        for row in rows:
            line = [col.value for col in row]
            #print(line)
            if (line[5] != None or line[6] != None) and (i > 2):
                ws['F' + str(i)] = None
                ws['G' + str(i)] = None
            if j < len(valueArray) and i > 2:
                addr, value = valueArray[j]
                print('addr=', addr, '  value=', value)
                ws['F' + str(i)] = addr
                ws['G' + str(i)] = value
                j += 1
            i += 1
        wb.save(ExcelFullName)

    except IOError:
        print('IOError: 没有找到 Excel 日志文件或写入文件失败！')
    else:
        print('读取的寄存器值已写入 Excel 日志文件成功！')


def writeExcel(excelName, sheet, column, rowStart, rowEnd, value):
    # 获取文件路径
    excelPath = os.path.join(os.getcwd())
    #print(excelPath)
    ExcelFullName = os.path.join(excelPath, excelName)
    print(ExcelFullName)

    try:
        wb = load_workbook(ExcelFullName)
        ws = wb[sheet]
        row = rowStart
        i = 0
        while 1:
            if row < rowEnd + 1:
                cell = column + str(row)
                ws[cell] = value[i]
                row += 1
                i += 1
            else:
                break
        wb.save(ExcelFullName)
    except IOError:
        print('IOError: 没有找到 Excel 日志文件或读取文件失败！')
    else:
        print('参数内容写入 Excel 日志文件成功！')


def initChart(excelName):
    excelPath = os.path.join(os.getcwd())
    #print(excelPath)
    ExcelFullName = os.path.join(excelPath, excelName)
    print(ExcelFullName)

    try:
        wb = load_workbook(ExcelFullName)

        ws = wb['InBand_Spur_test']

        chart1 = LineChart()
        chart1.type = "col"
        chart1.style = 12
        chart1.title = "InBand Spur"
        chart1.y_axis.title = 'Power Level'
        chart1.x_axis.title = 'Freq'
        data = Reference(ws, min_col=3, min_row=14, max_row=389, max_col=4)
        cats = Reference(ws, min_col=1, min_row=15, max_row=389)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.shape = 4
        ws.add_chart(chart1, "F16")

        ws1 = wb['Return_loss_Test']

        chart2 = LineChart()
        chart2.type = "col"
        chart2.style = 12
        chart2.title = "UP Band Return Loss"
        chart2.y_axis.title = 'dB'
        chart2.x_axis.title = 'MHz'
        data2 = Reference(ws1, min_col=3, min_row=5, max_row=66, max_col=4)
        cats2 = Reference(ws1, min_col=2, min_row=6, max_row=66)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        chart2.shape = 4
        ws1.add_chart(chart2, "G6")

        chart3 = LineChart()
        chart3.type = "col"
        chart3.style = 12
        chart3.title = "DS Band Return Loss"
        chart3.y_axis.title = 'dB'
        chart3.x_axis.title = 'MHz'
        data3 = Reference(ws1, min_col=3, min_row=70, max_row=803, max_col=4)
        cats3 = Reference(ws1, min_col=2, min_row=71, max_row=803)
        chart3.add_data(data3, titles_from_data=True)
        chart3.set_categories(cats3)
        chart3.shape = 4
        ws1.add_chart(chart3, "G68")

        wb.save(ExcelFullName)
    except IOError:
        print('IOError: 没有找到 Excel 日志文件或读取文件失败！')
    else:
        print('参数内容写入 Excel 日志文件成功！')


def initChartIPTuner(excelName, sheetName):
    excelPath = os.path.join(os.getcwd())
    # print(excelPath)
    ExcelFullName = os.path.join(excelPath, excelName)
    print(ExcelFullName)

    try:
        wb = load_workbook(ExcelFullName)

        ws = wb[sheetName]

        chart1 = LineChart()
        chart1.type = "col"
        chart1.style = 12
        chart1.title = "Tuner1 64QAM Sensitivity"
        chart1.y_axis.title = 'Power Level'
        chart1.x_axis.title = 'Freq'
        data = Reference(ws, min_col=2, min_row=16, max_row=110, max_col=5)
        cats = Reference(ws, min_col=1, min_row=17, max_row=110)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.shape = 4
        ws.add_chart(chart1, "M16")

        #ws1 = wb['report_tuner0']

        chart2 = LineChart()
        chart2.type = "col"
        chart2.style = 12
        chart2.title = "Tuner1 256QAM Sensitivity"
        chart2.y_axis.title = 'Power Level'
        chart2.x_axis.title = 'Freq'
        data2 = Reference(ws, min_col=8, min_row=16, max_row=110, max_col=11)
        cats2 = Reference(ws, min_col=7, min_row=15, max_row=110)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        chart2.shape = 4
        ws.add_chart(chart2, "M33")

        wb.save(ExcelFullName)
    except IOError:
        print('IOError: 没有找到 Excel 日志文件或读取文件失败！')
    else:
        print('参数内容写入 Excel 日志文件成功！')


if __name__ == '__main__':
    #str_to_hex('aabbccdd')
    #readExcel('FPGARegTest.xlsx')
    #writeExcel('IPTunerTest.xlsx', 'testResult', 'C', 17, 17, data1)
    #initChart('test.xlsx')
    #initChartIPTuner('IPTunerTest.xlsx', 'report_tuner0')
    #initChartIPTuner('IPTunerTest.xlsx', 'report_tuner1')
    writeAddrValueToExcel('FPGARegTest.xlsx', 'Sheet1', data2)